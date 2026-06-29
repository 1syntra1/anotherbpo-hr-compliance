import os
import io
import csv
import sqlite3
from datetime import datetime, date, timedelta
from functools import wraps
from flask import (
    Flask, g, render_template, request, redirect,
    url_for, flash, send_file, jsonify, session
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("HR_AUDIT_SECRET", "anotherbpo-hr-compliance-dev-key")

PURGE_HOLD_DAYS = 30  # admin-approved deletions are held this many days before purge

# ── Session behaviour ────────────────────────────────────────
INACTIVITY_TIMEOUT = 60  # seconds — auto-logout when "stay logged in" is NOT chosen
app.permanent_session_lifetime = timedelta(days=30)  # "stay logged in" duration


@app.before_request
def enforce_inactivity_timeout():
    """Sign out non-'remember' sessions after INACTIVITY_TIMEOUT of no requests.
    (A client-side idle timer logs the user out even without a request; this is the
    server-side backstop.) Sessions with 'stay logged in' are exempt."""
    if not session.get("logged_in"):
        return
    if session.get("remember"):
        return
    if request.endpoint in ("logout", "static"):
        return
    now = datetime.now().timestamp()
    last = session.get("last_activity")
    if last is not None and (now - last) > INACTIVITY_TIMEOUT:
        session.clear()
        flash("You were signed out after 1 minute of inactivity.", "info")
        return redirect(url_for("login"))
    session["last_activity"] = now


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("That area is restricted to administrators.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

DATABASE = os.path.join(os.path.dirname(__file__), "data", "hr_audit.db")
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "data", "uploads")
ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}

os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ─── Database helpers ────────────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE, detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DATABASE, detect_types=sqlite3.PARSE_DECLTYPES)
    db.row_factory = sqlite3.Row
    db.executescript("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_number TEXT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            id_number TEXT,
            race TEXT,
            gender TEXT,
            disability TEXT DEFAULT 'No',
            occupational_level TEXT,
            department TEXT,
            job_title TEXT,
            salary REAL,
            employment_date TEXT,
            province TEXT,
            nationality TEXT,
            coida_class_code TEXT,
            coida_subclass TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS training_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            employee_name TEXT,
            intervention_type TEXT,
            nqf_level TEXT,
            cost REAL,
            provider TEXT,
            target_date TEXT,
            actual_date TEXT,
            certificate TEXT,
            status TEXT DEFAULT 'Planned',
            seta TEXT,
            sdl_number TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (employee_id) REFERENCES employees(id)
        );

        CREATE TABLE IF NOT EXISTS coida_classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_code TEXT,
            subclass TEXT,
            description TEXT,
            num_employees INTEGER DEFAULT 0,
            annual_earnings REAL DEFAULT 0,
            assessment_year TEXT,
            employer_reg_number TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS logs_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            insurer TEXT,
            policy_number TEXT,
            issue_date TEXT,
            expiry_date TEXT,
            status TEXT DEFAULT 'Valid',
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS bbbee_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scorecard_type TEXT,
            ownership_score REAL DEFAULT 0,
            management_score REAL DEFAULT 0,
            skills_score REAL DEFAULT 0,
            esd_score REAL DEFAULT 0,
            sed_score REAL DEFAULT 0,
            total_score REAL DEFAULT 0,
            bbbee_level INTEGER DEFAULT 8,
            black_owned_percent REAL DEFAULT 0,
            assessment_year TEXT,
            company_name TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS company_info (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT,
            reg_number TEXT,
            reporting_year TEXT,
            designated_employer TEXT,
            seta TEXT,
            sdl_number TEXT,
            skills_levy REAL DEFAULT 0,
            employer_reg_number TEXT,
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS checklist_state (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            module TEXT NOT NULL,
            item_key TEXT NOT NULL,
            checked INTEGER DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now')),
            UNIQUE(module, item_key)
        );

        CREATE TABLE IF NOT EXISTS custom_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            module TEXT NOT NULL,
            step_key TEXT NOT NULL,
            text TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS hidden_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            module TEXT NOT NULL,
            item_key TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(module, item_key)
        );

        CREATE TABLE IF NOT EXISTS gbs_claim (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT,
            claim_quarter TEXT,
            claim_year TEXT,
            dtic_notification_date TEXT,
            submission_deadline TEXT,
            num_campaigns INTEGER DEFAULT 0,
            num_qualifying_employees INTEGER DEFAULT 0,
            export_revenue REAL DEFAULT 0,
            premises TEXT,
            auditor TEXT,
            notes TEXT,
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS gbs_projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quarter TEXT NOT NULL,                     -- 'Q1' / 'Q2' / 'Q3' / 'Q4'
            year INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'Draft',      -- Draft / In Progress / Submitted
            company_name TEXT,
            dtic_notification_date TEXT,
            submission_deadline TEXT,
            num_campaigns INTEGER DEFAULT 0,
            num_qualifying_employees INTEGER DEFAULT 0,
            export_revenue REAL DEFAULT 0,
            premises TEXT,
            auditor TEXT,
            notes TEXT,
            submitted_at TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',         -- 'admin' or 'user'
            full_name TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')),
            last_login TEXT
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT DEFAULT (datetime('now')),
            username TEXT,
            role TEXT,
            action TEXT,
            category TEXT,                             -- auth / project / task / trash / user / deadline
            detail TEXT,
            target TEXT
        );

        CREATE TABLE IF NOT EXISTS deadlines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dkey TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            authority TEXT,
            due_date TEXT,                             -- ISO date, or NULL for "varies"
            note TEXT,
            sort_order INTEGER DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now'))
        );
    """)

    # ── Migrations: add trash/soft-delete columns to gbs_projects ──
    existing = [r["name"] for r in db.execute("PRAGMA table_info(gbs_projects)").fetchall()]
    migrations = {
        "trashed": "INTEGER DEFAULT 0",
        "deleted_at": "TEXT",
        "deleted_by": "TEXT",
        "purge_at": "TEXT",
        "purge_approved_at": "TEXT",
        "purge_approved_by": "TEXT",
    }
    for col, coldef in migrations.items():
        if col not in existing:
            db.execute("ALTER TABLE gbs_projects ADD COLUMN %s %s" % (col, coldef))

    # ── Seed default accounts if none exist ──
    count = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
    if count == 0:
        db.execute(
            "INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)",
            ("admin", generate_password_hash("anotherbpo2024"), "admin", "System Administrator"),
        )
        db.execute(
            "INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)",
            ("user", generate_password_hash("anotherbpo2024"), "user", "Standard User"),
        )

    # ── Seed submission deadlines if none exist ──
    dcount = db.execute("SELECT COUNT(*) AS c FROM deadlines").fetchone()["c"]
    if dcount == 0:
        seed_deadlines = [
            # dkey, title, authority, due_date, note, sort_order
            ("wsp", "WSP & ATR", "SETA", "2026-06-30", "Workplace Skills Plan & Annual Training Report", 1),
            ("cipc_bo", "CIPC Beneficial Ownership", "CIPC", "2026-07-30", "Annual beneficial ownership filing", 2),
            ("ee", "Employment Equity (EEA2 & EEA4)", "DoEL", "2027-01-15", "Online submission deadline", 3),
            ("coida", "COIDA Return of Earnings (W.As.8)", "Compensation Fund", "2027-03-31", "Annual Return of Earnings", 4),
            ("gbs", "GBS Incentive Claim", "the dtic", None, "Within 30 days of dtic notification", 5),
            ("bbbee", "BBBEE Verification", "SANAS-accredited agency", None, "Annual (date varies)", 6),
        ]
        db.executemany(
            "INSERT INTO deadlines (dkey, title, authority, due_date, note, sort_order) VALUES (?,?,?,?,?,?)",
            seed_deadlines,
        )

    db.commit()
    db.close()


# Initialise DB on startup
with app.app_context():
    init_db()


RACES = ["African", "Coloured", "Indian", "White", "Foreign National"]
GENDERS = ["Male", "Female"]
OCC_LEVELS = [
    "Top Management", "Senior Management", "Professionally Qualified",
    "Skilled Technical", "Semi-Skilled", "Unskilled", "Non-Permanent"
]
PROVINCES = [
    "Gauteng", "Western Cape", "KwaZulu-Natal", "Eastern Cape",
    "Limpopo", "Mpumalanga", "North West", "Free State", "Northern Cape"
]
SETAS = [
    "AGRISETA", "BANKSETA", "CATHSSETA", "CHIETA", "CETA",
    "ETDP SETA", "FASSET", "FOODBEV", "HWSETA", "INSETA",
    "LGSETA", "MAPPP SETA", "MerSETA", "MICT SETA", "MQA",
    "NSDMS", "PSETA", "SASSETA", "Services SETA", "TETA", "W&RSETA"
]


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── Audit log ───────────────────────────────────────────────────────────────

def log_action(action, category, detail, target=None, username=None):
    """Write an entry to the audit log. Never raises into the request flow."""
    try:
        db = get_db()
        db.execute(
            "INSERT INTO audit_log (username, role, action, category, detail, target) VALUES (?,?,?,?,?,?)",
            (
                username or session.get("username", "unknown"),
                session.get("role", "-"),
                action, category, detail, target,
            ),
        )
        db.commit()
    except Exception:
        pass


def friendly_scope(db, module):
    """Turn a checklist scope like 'gbs:5' into a readable label."""
    if module == "ee":
        return "Employment Equity"
    if module and module.startswith("gbs:"):
        try:
            pid = int(module.split(":", 1)[1])
            row = db.execute("SELECT quarter, year FROM gbs_projects WHERE id=?", (pid,)).fetchone()
            if row:
                return "GBS %s %s" % (row["quarter"], row["year"])
        except (ValueError, IndexError):
            pass
    return module


def resolve_item_text(db, module, item_key):
    """Best-effort human-readable text for a checklist item key."""
    if item_key and item_key.startswith("custom-"):
        try:
            cid = int(item_key.split("-", 1)[1])
            row = db.execute("SELECT text FROM custom_tasks WHERE id=?", (cid,)).fetchone()
            if row:
                return row["text"]
        except (ValueError, IndexError):
            pass
        return "(custom task)"
    # Predefined: search the checklist definition for the base module
    from modules.checklists import get_checklist
    base = module.split(":", 1)[0] if module else module
    checklist = get_checklist(base)
    if checklist:
        for step in checklist["steps"]:
            for it in step["items"]:
                if it["key"] == item_key:
                    return it["text"]
    return item_key


# ─── Trash / purge helpers ───────────────────────────────────────────────────

def purge_expired_projects(db):
    """Permanently delete trashed projects whose 30-day hold has elapsed."""
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    rows = db.execute(
        "SELECT * FROM gbs_projects WHERE trashed=1 AND purge_at IS NOT NULL AND purge_at <= ?",
        (now,),
    ).fetchall()
    for r in rows:
        scope = "gbs:%d" % r["id"]
        db.execute("DELETE FROM gbs_projects WHERE id=?", (r["id"],))
        db.execute("DELETE FROM checklist_state WHERE module=?", (scope,))
        db.execute("DELETE FROM custom_tasks WHERE module=?", (scope,))
        db.execute("DELETE FROM hidden_tasks WHERE module=?", (scope,))
        log_action(
            "trash_purge", "trash",
            "Auto-purged after %d-day hold: GBS %s %s" % (PURGE_HOLD_DAYS, r["quarter"], r["year"]),
            target="gbs:%d" % r["id"], username="system",
        )
    if rows:
        db.commit()
    return len(rows)


# ─── Checklist helpers ───────────────────────────────────────────────────────

def get_checklist_states(db, module):
    """Return {item_key: bool} for all stored checkbox states of a module."""
    rows = db.execute(
        "SELECT item_key, checked FROM checklist_state WHERE module=?", (module,)
    ).fetchall()
    return {r["item_key"]: bool(r["checked"]) for r in rows}


def get_custom_tasks(db, module):
    """Return {step_key: [ {id, text, key} ]} for a module's user-added tasks."""
    rows = db.execute(
        "SELECT id, step_key, text FROM custom_tasks WHERE module=? ORDER BY id",
        (module,),
    ).fetchall()
    grouped = {}
    for r in rows:
        grouped.setdefault(r["step_key"], []).append(
            {"id": r["id"], "text": r["text"], "key": "custom-%d" % r["id"]}
        )
    return grouped


def get_hidden_tasks(db, module):
    """Return a set of item_keys hidden for this module."""
    rows = db.execute(
        "SELECT item_key FROM hidden_tasks WHERE module=?", (module,)
    ).fetchall()
    return {r["item_key"] for r in rows}


def checklist_progress(checklist, states, custom_by_step=None, hidden=None):
    """Return (done, total, percent) including predefined + custom tasks,
    excluding any items in `hidden`."""
    hidden = hidden or set()
    valid_keys = {
        it["key"] for step in checklist["steps"] for it in step["items"]
        if it["key"] not in hidden
    }
    if custom_by_step:
        for tasks in custom_by_step.values():
            for t in tasks:
                if t["key"] not in hidden:
                    valid_keys.add(t["key"])
    total = len(valid_keys)
    done = sum(1 for k in valid_keys if states.get(k))
    pct = round(done / total * 100) if total else 0
    return done, total, pct


# ─── Landing / Auth ──────────────────────────────────────────────────────────

@app.route("/home")
def landing():
    return render_template("landing.html", now=datetime.now)


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE username=? AND active=1", (username,)
        ).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            remember = bool(request.form.get("remember"))
            session["logged_in"] = True
            session["username"] = user["username"]
            session["role"] = user["role"]
            session["full_name"] = user["full_name"] or user["username"]
            session["remember"] = remember
            session.permanent = remember  # remember = 30-day cookie; else session cookie
            session["last_activity"] = datetime.now().timestamp()
            db.execute(
                "UPDATE users SET last_login=datetime('now') WHERE id=?", (user["id"],)
            )
            db.commit()
            log_action("login", "auth",
                       "Signed in (%s%s)" % (user["role"], ", stay logged in" if remember else ""))
            flash(f"Welcome back, {session['full_name']}!", "success")
            return redirect(url_for("dashboard"))
        log_action("login_failed", "auth",
                   "Failed sign-in attempt", target=username, username=username or "unknown")
        flash("Invalid username or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    timed_out = request.args.get("timeout")
    if session.get("logged_in"):
        log_action("logout", "auth",
                   "Auto sign-out (1 min inactivity)" if timed_out else "Signed out")
    session.clear()
    if timed_out:
        flash("You were signed out after 1 minute of inactivity.", "info")
        return redirect(url_for("login"))
    flash("You have been logged out.", "info")
    return redirect(url_for("landing"))


# ─── Deadlines ───────────────────────────────────────────────────────────────

def _deadline_view(row):
    """Augment a deadline row with display string, days remaining, and a status class."""
    d = dict(row)
    due = d.get("due_date")
    if due:
        try:
            dd = date.fromisoformat(due)
            days = (dd - date.today()).days
            d["due_display"] = dd.strftime("%d %b %Y")
            d["days"] = days
            if days < 0:
                d["status_label"] = "Overdue by %d day%s" % (abs(days), "" if abs(days) == 1 else "s")
                d["status_class"] = "red"
            elif days == 0:
                d["status_label"] = "Due today"
                d["status_class"] = "red"
            elif days <= 14:
                d["status_label"] = "In %d day%s" % (days, "" if days == 1 else "s")
                d["status_class"] = "red"
            elif days <= 45:
                d["status_label"] = "In %d days" % days
                d["status_class"] = "amber"
            else:
                d["status_label"] = "In %d days" % days
                d["status_class"] = "teal"
            return d
        except ValueError:
            pass
    # No (valid) date → "varies"
    d["due_display"] = d.get("note") or "Varies"
    d["days"] = None
    d["status_label"] = "Varies"
    d["status_class"] = "grey"
    return d


def get_deadlines(db):
    rows = db.execute("SELECT * FROM deadlines").fetchall()
    items = [_deadline_view(r) for r in rows]
    # Soonest dated first; undated ("varies") last; tie-break on sort_order
    items.sort(key=lambda x: (x["days"] is None, x["days"] if x["days"] is not None else 0, x["sort_order"]))
    return items


# ─── Dashboard ───────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    db = get_db()
    emp_count = db.execute("SELECT COUNT(*) as c FROM employees").fetchone()["c"]
    training_count = db.execute("SELECT COUNT(*) as c FROM training_records").fetchone()["c"]
    logs_count = db.execute("SELECT COUNT(*) as c FROM logs_records").fetchone()["c"]
    bbbee_row = db.execute(
        "SELECT * FROM bbbee_data ORDER BY id DESC LIMIT 1"
    ).fetchone()
    deadlines = get_deadlines(db)
    return render_template(
        "dashboard.html",
        emp_count=emp_count,
        training_count=training_count,
        logs_count=logs_count,
        bbbee_row=bbbee_row,
        deadlines=deadlines,
        current_year=datetime.now().year,
    )


@app.route("/deadlines/<int:deadline_id>/update", methods=["POST"])
@login_required
def deadline_update(deadline_id):
    db = get_db()
    row = db.execute("SELECT * FROM deadlines WHERE id=?", (deadline_id,)).fetchone()
    if not row:
        flash("Deadline not found.", "danger")
        return redirect(url_for("dashboard"))
    due_date = (request.form.get("due_date") or "").strip() or None
    note = (request.form.get("note") or "").strip()
    # Validate date if supplied
    if due_date:
        try:
            date.fromisoformat(due_date)
        except ValueError:
            flash("Invalid date. Please use the date picker.", "danger")
            return redirect(url_for("dashboard"))
    db.execute(
        "UPDATE deadlines SET due_date=?, note=?, updated_at=datetime('now') WHERE id=?",
        (due_date, note, deadline_id),
    )
    db.commit()
    log_action("deadline_update", "deadline",
               "Set %s due date to %s" % (row["title"], due_date or "(varies)"),
               target=row["dkey"])
    flash("Deadline updated.", "success")
    return redirect(url_for("dashboard"))


# ─── Employees ───────────────────────────────────────────────────────────────

@app.route("/employees", methods=["GET", "POST"])
@login_required
def employees():
    db = get_db()
    if request.method == "POST":
        data = request.form
        db.execute(
            """INSERT INTO employees
               (emp_number, first_name, last_name, id_number, race, gender,
                disability, occupational_level, department, job_title, salary,
                employment_date, province, nationality, coida_class_code, coida_subclass)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                data.get("emp_number"), data.get("first_name"), data.get("last_name"),
                data.get("id_number"), data.get("race"), data.get("gender"),
                data.get("disability", "No"), data.get("occupational_level"),
                data.get("department"), data.get("job_title"),
                data.get("salary") or None, data.get("employment_date"),
                data.get("province"), data.get("nationality"),
                data.get("coida_class_code"), data.get("coida_subclass"),
            ),
        )
        db.commit()
        flash("Employee added successfully.", "success")
        return redirect(url_for("employees"))

    all_employees = db.execute(
        "SELECT * FROM employees ORDER BY last_name, first_name"
    ).fetchall()
    return render_template(
        "employees.html",
        employees=all_employees,
        races=RACES,
        genders=GENDERS,
        occ_levels=OCC_LEVELS,
        provinces=PROVINCES,
    )


@app.route("/employees/upload", methods=["POST"])
@login_required
def employees_upload():
    db = get_db()
    if "file" not in request.files:
        flash("No file selected.", "danger")
        return redirect(url_for("employees"))
    f = request.files["file"]
    if f.filename == "" or not allowed_file(f.filename):
        flash("Invalid file. Use CSV or Excel.", "danger")
        return redirect(url_for("employees"))

    filename = secure_filename(f.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(filepath)

    rows_added = 0
    try:
        ext = filename.rsplit(".", 1)[1].lower()
        if ext == "csv":
            with open(filepath, newline="", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    _insert_employee_row(db, row)
                    rows_added += 1
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rowdict = dict(zip(headers, row))
                _insert_employee_row(db, rowdict)
                rows_added += 1
        db.commit()
        flash(f"{rows_added} employee(s) imported successfully.", "success")
    except Exception as e:
        flash(f"Import error: {e}", "danger")

    return redirect(url_for("employees"))


def _insert_employee_row(db, row):
    db.execute(
        """INSERT INTO employees
           (emp_number, first_name, last_name, id_number, race, gender,
            disability, occupational_level, department, job_title, salary,
            employment_date, province, nationality, coida_class_code, coida_subclass)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            row.get("emp_number") or row.get("Emp Number"),
            row.get("first_name") or row.get("First Name"),
            row.get("last_name") or row.get("Last Name"),
            row.get("id_number") or row.get("ID Number"),
            row.get("race") or row.get("Race"),
            row.get("gender") or row.get("Gender"),
            row.get("disability") or row.get("Disability") or "No",
            row.get("occupational_level") or row.get("Occupational Level"),
            row.get("department") or row.get("Department"),
            row.get("job_title") or row.get("Job Title"),
            row.get("salary") or row.get("Salary"),
            row.get("employment_date") or row.get("Employment Date"),
            row.get("province") or row.get("Province"),
            row.get("nationality") or row.get("Nationality"),
            row.get("coida_class_code") or row.get("COIDA Class"),
            row.get("coida_subclass") or row.get("COIDA Subclass"),
        ),
    )


@app.route("/employees/<int:emp_id>/edit", methods=["GET", "POST"])
@login_required
def edit_employee(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if emp is None:
        flash("Employee not found.", "danger")
        return redirect(url_for("employees"))

    if request.method == "POST":
        data = request.form
        db.execute(
            """UPDATE employees SET
               emp_number=?, first_name=?, last_name=?, id_number=?, race=?,
               gender=?, disability=?, occupational_level=?, department=?,
               job_title=?, salary=?, employment_date=?, province=?,
               nationality=?, coida_class_code=?, coida_subclass=?
               WHERE id=?""",
            (
                data.get("emp_number"), data.get("first_name"), data.get("last_name"),
                data.get("id_number"), data.get("race"), data.get("gender"),
                data.get("disability", "No"), data.get("occupational_level"),
                data.get("department"), data.get("job_title"),
                data.get("salary") or None, data.get("employment_date"),
                data.get("province"), data.get("nationality"),
                data.get("coida_class_code"), data.get("coida_subclass"),
                emp_id,
            ),
        )
        db.commit()
        flash("Employee updated.", "success")
        return redirect(url_for("employees"))

    return render_template(
        "employees.html",
        edit_emp=emp,
        employees=db.execute("SELECT * FROM employees ORDER BY last_name").fetchall(),
        races=RACES,
        genders=GENDERS,
        occ_levels=OCC_LEVELS,
        provinces=PROVINCES,
    )


@app.route("/employees/<int:emp_id>/delete", methods=["POST"])
@login_required
def delete_employee(emp_id):
    db = get_db()
    db.execute("DELETE FROM employees WHERE id=?", (emp_id,))
    db.commit()
    flash("Employee deleted.", "warning")
    return redirect(url_for("employees"))


@app.route("/employees/template")
@login_required
def employee_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "emp_number", "first_name", "last_name", "id_number", "race", "gender",
        "disability", "occupational_level", "department", "job_title", "salary",
        "employment_date", "province", "nationality", "coida_class_code", "coida_subclass"
    ])
    writer.writerow([
        "EMP001", "John", "Doe", "8001015009087", "African", "Male",
        "No", "Skilled Technical", "IT", "Developer", "25000",
        "2020-01-15", "Gauteng", "South African", "A", "A1"
    ])
    output.seek(0)
    return send_file(
        io.BytesIO(output.read().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name="employee_template.csv",
    )


# ─── EE — Employment Equity (EEA2 & EEA4) ────────────────────────────────────

def _build_ee_matrix(employees):
    races_short = ["African", "Coloured", "Indian", "White", "Foreign National"]
    matrix = {}
    for lvl in OCC_LEVELS:
        matrix[lvl] = {r: {"M": 0, "F": 0} for r in races_short}
        matrix[lvl]["Total"] = 0
    for emp in employees:
        lvl = emp["occupational_level"]
        race = emp["race"]
        gender = emp["gender"]
        if lvl in matrix and race in races_short:
            g_key = "M" if gender == "Male" else "F"
            matrix[lvl][race][g_key] += 1
            matrix[lvl]["Total"] += 1
    return matrix


@app.route("/ee", methods=["GET", "POST"])
@login_required
def ee():
    db = get_db()
    if request.method == "POST":
        data = request.form
        info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
        if info:
            db.execute(
                """UPDATE company_info SET company_name=?, reg_number=?,
                   reporting_year=?, designated_employer=?, updated_at=datetime('now') WHERE id=?""",
                (data.get("company_name"), data.get("reg_number"),
                 data.get("reporting_year"), data.get("designated_employer"), info["id"])
            )
        else:
            db.execute(
                """INSERT INTO company_info (company_name, reg_number, reporting_year, designated_employer)
                   VALUES (?,?,?,?)""",
                (data.get("company_name"), data.get("reg_number"),
                 data.get("reporting_year"), data.get("designated_employer"))
            )
        db.commit()
        log_action("ee_update", "project", "Updated Employment Equity company details", target="ee")
        flash("Employment Equity company info saved.", "success")
        return redirect(url_for("ee"))

    employees = db.execute("SELECT * FROM employees").fetchall()
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    matrix = _build_ee_matrix(employees)
    disability_count = sum(1 for e in employees if e["disability"] == "Yes")

    from modules.checklists import get_checklist
    checklist = get_checklist("ee")
    states = get_checklist_states(db, "ee")
    custom = get_custom_tasks(db, "ee")
    hidden = get_hidden_tasks(db, "ee")
    done, total, pct = checklist_progress(checklist, states, custom, hidden)

    return render_template(
        "ee.html",
        matrix=matrix,
        occ_levels=OCC_LEVELS,
        disability_count=disability_count,
        company_info=company_info,
        total_employees=len(employees),
        checklist=checklist,
        checklist_states=states,
        custom_tasks=custom,
        hidden_tasks=hidden,
        cl_done=done,
        cl_total=total,
        cl_pct=pct,
    )


@app.route("/ee/generate", methods=["POST"])
@login_required
def ee_generate():
    from modules.gbs import generate_eea2, generate_eea4
    db = get_db()
    report_type = request.form.get("report_type", "eea2")
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    ci = dict(company_info) if company_info else {}

    if report_type == "eea2":
        data = generate_eea2(db, ci)
        fname = "EEA2_Report.xlsx"
    else:
        data = generate_eea4(db, ci)
        fname = "EEA4_Report.xlsx"

    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ─── GBS — Global Business Services Quarterly Claim Projects ─────────────────

QUARTERS = [
    ("Q1", "Q1 — 1 January"),
    ("Q2", "Q2 — 1 April"),
    ("Q3", "Q3 — 1 July"),
    ("Q4", "Q4 — 1 October"),
]


def _gbs_scope(project_id):
    """Checklist 'module' key for a specific GBS project."""
    return "gbs:%d" % int(project_id)


def _gbs_project_progress(db, project_id):
    """Return (done, total, percent) for a project's checklist."""
    from modules.checklists import get_checklist
    checklist = get_checklist("gbs")
    if not checklist:
        return 0, 0, 0
    scope = _gbs_scope(project_id)
    states = get_checklist_states(db, scope)
    custom = get_custom_tasks(db, scope)
    hidden = get_hidden_tasks(db, scope)
    return checklist_progress(checklist, states, custom, hidden)


@app.route("/gbs", methods=["GET"])
@login_required
def gbs():
    """Projects dashboard — all (non-trashed) quarterly claims grouped by year."""
    db = get_db()
    purge_expired_projects(db)  # housekeeping: clear any holds that have elapsed

    rows = db.execute(
        "SELECT * FROM gbs_projects WHERE trashed=0 ORDER BY year DESC, quarter DESC, id DESC"
    ).fetchall()

    # Group by year
    by_year = {}
    for r in rows:
        proj = dict(r)
        done, total, pct = _gbs_project_progress(db, proj["id"])
        proj["cl_done"] = done
        proj["cl_total"] = total
        proj["cl_pct"] = pct
        by_year.setdefault(proj["year"], []).append(proj)

    # Sort quarters within each year (Q4 down to Q1 for newest first)
    quarter_order = {"Q4": 4, "Q3": 3, "Q2": 2, "Q1": 1}
    for yr in by_year:
        by_year[yr].sort(key=lambda p: quarter_order.get(p["quarter"], 0), reverse=True)
    grouped = sorted(by_year.items(), reverse=True)  # newest year first

    # Stats
    total_projects = len(rows)
    submitted = sum(1 for r in rows if r["status"] == "Submitted")
    in_progress = sum(1 for r in rows if r["status"] == "In Progress")
    drafts = sum(1 for r in rows if r["status"] == "Draft")
    trash_count = db.execute(
        "SELECT COUNT(*) AS c FROM gbs_projects WHERE trashed=1"
    ).fetchone()["c"]

    now = datetime.now()
    current_q = "Q%d" % ((now.month - 1) // 3 + 1)

    return render_template(
        "gbs_projects.html",
        grouped=grouped,
        total_projects=total_projects,
        submitted=submitted,
        in_progress=in_progress,
        drafts=drafts,
        trash_count=trash_count,
        current_year=now.year,
        current_quarter=current_q,
        quarters=QUARTERS,
    )


@app.route("/gbs/new", methods=["POST"])
@login_required
def gbs_new():
    db = get_db()
    quarter = (request.form.get("quarter") or "").strip()
    year = request.form.get("year")
    if quarter not in {q[0] for q in QUARTERS}:
        flash("Please select a valid quarter.", "danger")
        return redirect(url_for("gbs"))
    try:
        year = int(year)
    except (TypeError, ValueError):
        flash("Please provide a valid year.", "danger")
        return redirect(url_for("gbs"))

    # Warn if a non-trashed project already exists for this quarter+year (but allow it)
    existing = db.execute(
        "SELECT id FROM gbs_projects WHERE quarter=? AND year=? AND trashed=0", (quarter, year)
    ).fetchone()
    if existing and not request.form.get("force"):
        flash(
            f"A project for {quarter} {year} already exists. Open it from the list, or tick 'Create anyway' to make a second one.",
            "warning",
        )
        return redirect(url_for("gbs"))

    cur = db.execute(
        "INSERT INTO gbs_projects (quarter, year, status) VALUES (?, ?, 'Draft')",
        (quarter, year),
    )
    db.commit()
    log_action("project_create", "project",
               f"Created GBS quarterly claim {quarter} {year}", target=f"gbs:{cur.lastrowid}")
    flash(f"New GBS project created: {quarter} {year}.", "success")
    return redirect(url_for("gbs_project", project_id=cur.lastrowid))


@app.route("/gbs/<int:project_id>", methods=["GET", "POST"])
@login_required
def gbs_project(project_id):
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    if not proj:
        flash("Project not found.", "danger")
        return redirect(url_for("gbs"))
    if proj["trashed"]:
        flash("That project is in the trash. An administrator must restore it first.", "warning")
        return redirect(url_for("gbs"))

    readonly = proj["status"] == "Submitted"

    if request.method == "POST":
        if readonly:
            flash("This project has been submitted. Reopen it to edit.", "warning")
            return redirect(url_for("gbs_project", project_id=project_id))
        data = request.form
        # Auto-promote Draft → In Progress on first save
        new_status = proj["status"]
        if new_status == "Draft":
            new_status = "In Progress"
        db.execute(
            """UPDATE gbs_projects SET company_name=?, dtic_notification_date=?,
               submission_deadline=?, num_campaigns=?, num_qualifying_employees=?,
               export_revenue=?, premises=?, auditor=?, notes=?, status=?,
               updated_at=datetime('now') WHERE id=?""",
            (
                data.get("company_name"), data.get("dtic_notification_date"),
                data.get("submission_deadline"),
                data.get("num_campaigns") or 0, data.get("num_qualifying_employees") or 0,
                data.get("export_revenue") or 0, data.get("premises"),
                data.get("auditor"), data.get("notes"), new_status, project_id,
            ),
        )
        db.commit()
        log_action("project_update", "project",
                   f"Updated claim details for GBS {proj['quarter']} {proj['year']}",
                   target=f"gbs:{project_id}")
        flash("Claim details saved.", "success")
        return redirect(url_for("gbs_project", project_id=project_id))

    from modules.checklists import get_checklist
    checklist = get_checklist("gbs")
    scope = _gbs_scope(project_id)
    states = get_checklist_states(db, scope)
    custom = get_custom_tasks(db, scope)
    hidden = get_hidden_tasks(db, scope)
    done, total, pct = checklist_progress(checklist, states, custom, hidden)

    return render_template(
        "gbs_project.html",
        project=proj,
        scope=scope,
        readonly=readonly,
        checklist=checklist,
        checklist_states=states,
        custom_tasks=custom,
        hidden_tasks=hidden,
        cl_done=done,
        cl_total=total,
        cl_pct=pct,
        current_year=datetime.now().year,
    )


@app.route("/gbs/<int:project_id>/submit", methods=["POST"])
@login_required
def gbs_project_submit(project_id):
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    if not proj:
        flash("Project not found.", "danger")
        return redirect(url_for("gbs"))
    db.execute(
        "UPDATE gbs_projects SET status='Submitted', submitted_at=datetime('now'), updated_at=datetime('now') WHERE id=?",
        (project_id,),
    )
    db.commit()
    log_action("project_submit", "project",
               f"Submitted & closed GBS {proj['quarter']} {proj['year']}", target=f"gbs:{project_id}")
    flash(f"Project {proj['quarter']} {proj['year']} submitted and closed.", "success")
    return redirect(url_for("gbs"))


@app.route("/gbs/<int:project_id>/reopen", methods=["POST"])
@login_required
def gbs_project_reopen(project_id):
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    db.execute(
        "UPDATE gbs_projects SET status='In Progress', submitted_at=NULL, updated_at=datetime('now') WHERE id=?",
        (project_id,),
    )
    db.commit()
    if proj:
        log_action("project_reopen", "project",
                   f"Reopened GBS {proj['quarter']} {proj['year']}", target=f"gbs:{project_id}")
    flash("Project reopened for editing.", "info")
    return redirect(url_for("gbs_project", project_id=project_id))


@app.route("/gbs/<int:project_id>/delete", methods=["POST"])
@login_required
def gbs_project_delete(project_id):
    """Soft-delete: move the project to the trash bin (recoverable by an admin)."""
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    if not proj:
        flash("Project not found.", "danger")
        return redirect(url_for("gbs"))
    db.execute(
        """UPDATE gbs_projects SET trashed=1, deleted_at=datetime('now'), deleted_by=?,
           purge_at=NULL, purge_approved_at=NULL, purge_approved_by=NULL,
           updated_at=datetime('now') WHERE id=?""",
        (session.get("username", "unknown"), project_id),
    )
    db.commit()
    log_action("project_delete", "project",
               f"Moved GBS {proj['quarter']} {proj['year']} to trash", target=f"gbs:{project_id}")
    flash(
        f"Project {proj['quarter']} {proj['year']} moved to the trash bin. "
        "An administrator can restore it or approve permanent deletion.",
        "warning",
    )
    return redirect(url_for("gbs"))


# ─── Trash bin (admin only) ──────────────────────────────────────────────────

@app.route("/trash", methods=["GET"])
@admin_required
def trash():
    db = get_db()
    purge_expired_projects(db)

    rows = db.execute(
        "SELECT * FROM gbs_projects WHERE trashed=1 ORDER BY deleted_at DESC"
    ).fetchall()

    awaiting, pending = [], []
    now = datetime.now()
    for r in rows:
        p = dict(r)
        done, total, pct = _gbs_project_progress(db, p["id"])
        p["cl_done"], p["cl_total"], p["cl_pct"] = done, total, pct
        if p["purge_at"]:
            try:
                import math
                purge_dt = datetime.fromisoformat(p["purge_at"])
                secs = (purge_dt - now).total_seconds()
                p["days_left"] = max(0, int(math.ceil(secs / 86400.0)))
                p["purge_date_str"] = purge_dt.strftime("%d %b %Y")
            except ValueError:
                p["days_left"] = 0
                p["purge_date_str"] = p["purge_at"]
            pending.append(p)
        else:
            awaiting.append(p)

    return render_template(
        "trash.html",
        awaiting=awaiting,
        pending=pending,
        hold_days=PURGE_HOLD_DAYS,
    )


@app.route("/trash/<int:project_id>/restore", methods=["POST"])
@admin_required
def trash_restore(project_id):
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    if not proj:
        flash("Project not found.", "danger")
        return redirect(url_for("trash"))
    db.execute(
        """UPDATE gbs_projects SET trashed=0, deleted_at=NULL, deleted_by=NULL,
           purge_at=NULL, purge_approved_at=NULL, purge_approved_by=NULL,
           updated_at=datetime('now') WHERE id=?""",
        (project_id,),
    )
    db.commit()
    log_action("trash_restore", "trash",
               f"Restored GBS {proj['quarter']} {proj['year']} from trash", target=f"gbs:{project_id}")
    flash(f"Project {proj['quarter']} {proj['year']} restored.", "success")
    return redirect(url_for("trash"))


@app.route("/trash/<int:project_id>/approve", methods=["POST"])
@admin_required
def trash_approve(project_id):
    """Admin approves permanent deletion — starts the 30-day hold countdown."""
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    if not proj:
        flash("Project not found.", "danger")
        return redirect(url_for("trash"))
    purge_at = (datetime.now() + timedelta(days=PURGE_HOLD_DAYS)).isoformat(sep=" ", timespec="seconds")
    db.execute(
        """UPDATE gbs_projects SET purge_at=?, purge_approved_at=datetime('now'),
           purge_approved_by=?, updated_at=datetime('now') WHERE id=?""",
        (purge_at, session.get("username", "admin"), project_id),
    )
    db.commit()
    log_action("trash_approve_delete", "trash",
               f"Approved permanent deletion of GBS {proj['quarter']} {proj['year']} — "
               f"{PURGE_HOLD_DAYS}-day hold begins (purges {purge_at[:10]})",
               target=f"gbs:{project_id}")
    flash(
        f"Permanent deletion approved for {proj['quarter']} {proj['year']}. "
        f"It will be purged in {PURGE_HOLD_DAYS} days unless restored.",
        "warning",
    )
    return redirect(url_for("trash"))


@app.route("/trash/<int:project_id>/cancel", methods=["POST"])
@admin_required
def trash_cancel(project_id):
    """Cancel a pending purge — the item stays in trash, countdown removed."""
    db = get_db()
    proj = db.execute("SELECT * FROM gbs_projects WHERE id=?", (project_id,)).fetchone()
    if not proj:
        flash("Project not found.", "danger")
        return redirect(url_for("trash"))
    db.execute(
        """UPDATE gbs_projects SET purge_at=NULL, purge_approved_at=NULL,
           purge_approved_by=NULL, updated_at=datetime('now') WHERE id=?""",
        (project_id,),
    )
    db.commit()
    log_action("trash_cancel_delete", "trash",
               f"Cancelled the deletion hold on GBS {proj['quarter']} {proj['year']}",
               target=f"gbs:{project_id}")
    flash("Deletion hold cancelled. The project remains in the trash bin.", "info")
    return redirect(url_for("trash"))


# ─── WSP ─────────────────────────────────────────────────────────────────────

@app.route("/wsp", methods=["GET", "POST"])
@login_required
def wsp():
    db = get_db()
    if request.method == "POST":
        data = request.form
        info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
        if info:
            db.execute(
                "UPDATE company_info SET seta=?, sdl_number=?, skills_levy=?, updated_at=datetime('now') WHERE id=?",
                (data.get("seta"), data.get("sdl_number"), data.get("skills_levy") or 0, info["id"])
            )
        else:
            db.execute(
                "INSERT INTO company_info (seta, sdl_number, skills_levy) VALUES (?,?,?)",
                (data.get("seta"), data.get("sdl_number"), data.get("skills_levy") or 0)
            )
        db.commit()
        flash("WSP info saved.", "success")
        return redirect(url_for("wsp"))

    employees = db.execute("SELECT id, first_name, last_name FROM employees ORDER BY last_name").fetchall()
    training = db.execute("SELECT * FROM training_records ORDER BY created_at DESC").fetchall()
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    return render_template("wsp.html", employees=employees, training=training,
                           company_info=company_info, setas=SETAS)


@app.route("/wsp/training", methods=["POST"])
@login_required
def wsp_training():
    db = get_db()
    data = request.form
    emp_id = data.get("employee_id")
    emp_name = ""
    if emp_id:
        emp = db.execute("SELECT first_name, last_name FROM employees WHERE id=?", (emp_id,)).fetchone()
        if emp:
            emp_name = f"{emp['first_name']} {emp['last_name']}"
    db.execute(
        """INSERT INTO training_records
           (employee_id, employee_name, intervention_type, nqf_level, cost, provider,
            target_date, actual_date, certificate, status, seta, sdl_number)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            emp_id, emp_name, data.get("intervention_type"), data.get("nqf_level"),
            data.get("cost") or None, data.get("provider"), data.get("target_date"),
            data.get("actual_date"), data.get("certificate"),
            data.get("status", "Planned"), data.get("seta"), data.get("sdl_number"),
        ),
    )
    db.commit()
    flash("Training record added.", "success")
    return redirect(url_for("wsp"))


@app.route("/wsp/generate", methods=["POST"])
@login_required
def wsp_generate():
    from modules.wsp import generate_wsp, generate_atr, generate_wsp_word
    db = get_db()
    report_type = request.form.get("report_type", "wsp")
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    ci = dict(company_info) if company_info else {}

    if report_type == "wsp":
        data = generate_wsp(db, ci)
        return send_file(io.BytesIO(data),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="WSP_Report.xlsx")
    elif report_type == "atr":
        data = generate_atr(db, ci)
        return send_file(io.BytesIO(data),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="ATR_Report.xlsx")
    else:
        data = generate_wsp_word(db, ci)
        return send_file(io.BytesIO(data),
                         mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                         as_attachment=True, download_name="WSP_Report.docx")


# ─── BBBEE ───────────────────────────────────────────────────────────────────

@app.route("/bbbee", methods=["GET"])
@login_required
def bbbee():
    db = get_db()
    bbbee_row = db.execute("SELECT * FROM bbbee_data ORDER BY id DESC LIMIT 1").fetchone()
    return render_template("bbbee.html", bbbee_row=bbbee_row)


@app.route("/bbbee/save", methods=["POST"])
@login_required
def bbbee_save():
    from modules.bbbee import calculate_level
    db = get_db()
    data = request.form
    scorecard_type = data.get("scorecard_type", "Generic")
    ownership = float(data.get("ownership_score") or 0)
    management = float(data.get("management_score") or 0)
    skills = float(data.get("skills_score") or 0)
    esd = float(data.get("esd_score") or 0)
    sed = float(data.get("sed_score") or 0)
    total = ownership + management + skills + esd + sed
    black_pct = float(data.get("black_owned_percent") or 0)
    level, _ = calculate_level(scorecard_type, total, black_pct)

    db.execute(
        """INSERT INTO bbbee_data
           (scorecard_type, ownership_score, management_score, skills_score,
            esd_score, sed_score, total_score, bbbee_level, black_owned_percent,
            assessment_year, company_name)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (scorecard_type, ownership, management, skills, esd, sed, total, level,
         black_pct, data.get("assessment_year"), data.get("company_name"))
    )
    db.commit()
    flash(f"BBBEE scorecard saved. Level {level}.", "success")
    return redirect(url_for("bbbee"))


@app.route("/bbbee/generate", methods=["POST"])
@login_required
def bbbee_generate():
    from modules.bbbee import generate_scorecard_excel, generate_scorecard_pdf
    db = get_db()
    report_type = request.form.get("report_type", "excel")
    row = db.execute("SELECT * FROM bbbee_data ORDER BY id DESC LIMIT 1").fetchone()
    scorecard_data = dict(row) if row else {}

    if report_type == "pdf":
        data = generate_scorecard_pdf(scorecard_data)
        return send_file(io.BytesIO(data), mimetype="application/pdf",
                         as_attachment=True, download_name="BBBEE_Certificate.pdf")
    else:
        data = generate_scorecard_excel(scorecard_data)
        return send_file(io.BytesIO(data),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="BBBEE_Scorecard.xlsx")


# ─── COIDA ───────────────────────────────────────────────────────────────────

@app.route("/coida", methods=["GET"])
@login_required
def coida():
    db = get_db()
    earnings = db.execute("SELECT * FROM coida_classes ORDER BY class_code").fetchall()
    logs = db.execute("SELECT * FROM logs_records ORDER BY expiry_date").fetchall()
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    today = date.today().isoformat()

    # Update LOGS status dynamically
    updated_logs = []
    for log in logs:
        expiry = log["expiry_date"] or ""
        if expiry < today:
            status = "Expired"
        elif expiry <= date.today().replace(month=date.today().month).isoformat() or \
                (expiry[:7] <= date.today().isoformat()[:7]):
            # Expiring within 30 days
            from datetime import timedelta
            threshold = (date.today() + timedelta(days=30)).isoformat()
            status = "Expiring Soon" if expiry <= threshold else "Valid"
        else:
            status = "Valid"
        updated_logs.append({"log": log, "status": status})

    from modules.coida import calculate_assessment
    assessment = calculate_assessment(list(earnings))
    return render_template("coida.html", earnings=earnings, logs=updated_logs,
                           assessment=assessment, company_info=company_info,
                           today=today)


@app.route("/coida/earning", methods=["POST"])
@login_required
def coida_earning():
    db = get_db()
    data = request.form
    rec_id = data.get("record_id")
    if rec_id:
        db.execute(
            """UPDATE coida_classes SET class_code=?, subclass=?, description=?,
               num_employees=?, annual_earnings=?, assessment_year=?, employer_reg_number=?
               WHERE id=?""",
            (data.get("class_code"), data.get("subclass"), data.get("description"),
             data.get("num_employees") or 0, data.get("annual_earnings") or 0,
             data.get("assessment_year"), data.get("employer_reg_number"), rec_id)
        )
    else:
        db.execute(
            """INSERT INTO coida_classes
               (class_code, subclass, description, num_employees, annual_earnings,
                assessment_year, employer_reg_number)
               VALUES (?,?,?,?,?,?,?)""",
            (data.get("class_code"), data.get("subclass"), data.get("description"),
             data.get("num_employees") or 0, data.get("annual_earnings") or 0,
             data.get("assessment_year"), data.get("employer_reg_number"))
        )
    db.commit()
    flash("Earnings record saved.", "success")
    return redirect(url_for("coida"))


@app.route("/coida/logs", methods=["POST"])
@login_required
def coida_logs():
    db = get_db()
    data = request.form
    rec_id = data.get("record_id")
    if rec_id:
        db.execute(
            "UPDATE logs_records SET insurer=?, policy_number=?, issue_date=?, expiry_date=? WHERE id=?",
            (data.get("insurer"), data.get("policy_number"),
             data.get("issue_date"), data.get("expiry_date"), rec_id)
        )
    else:
        db.execute(
            "INSERT INTO logs_records (insurer, policy_number, issue_date, expiry_date) VALUES (?,?,?,?)",
            (data.get("insurer"), data.get("policy_number"),
             data.get("issue_date"), data.get("expiry_date"))
        )
    db.commit()
    flash("LOGS record saved.", "success")
    return redirect(url_for("coida"))


@app.route("/coida/generate", methods=["POST"])
@login_required
def coida_generate():
    from modules.coida import generate_was8
    db = get_db()
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    ci = dict(company_info) if company_info else {}
    ci["assessment_year"] = request.form.get("assessment_year", str(datetime.now().year))
    data = generate_was8(db, ci)
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="WAs8_Return.xlsx",
    )


# ─── API helpers ─────────────────────────────────────────────────────────────

# ─── Checklist API ───────────────────────────────────────────────────────────

def _base_module(scope):
    """A scope like 'gbs:5' uses the base checklist definition 'gbs'."""
    return scope.split(":", 1)[0] if scope else scope


@app.route("/checklist/toggle", methods=["POST"])
@login_required
def checklist_toggle():
    db = get_db()
    data = request.get_json(silent=True) or {}
    module = data.get("module", "")
    item_key = data.get("item_key", "")
    checked = 1 if data.get("checked") else 0
    if not module or not item_key:
        return jsonify({"ok": False, "error": "missing module or item_key"}), 400

    db.execute(
        """INSERT INTO checklist_state (module, item_key, checked, updated_at)
           VALUES (?, ?, ?, datetime('now'))
           ON CONFLICT(module, item_key)
           DO UPDATE SET checked=excluded.checked, updated_at=datetime('now')""",
        (module, item_key, checked),
    )
    db.commit()

    item_text = resolve_item_text(db, module, item_key)
    log_action(
        "task_toggle", "task",
        "%s task '%s' in %s" % (
            "Completed" if checked else "Un-checked", item_text, friendly_scope(db, module)),
        target=module,
    )

    from modules.checklists import get_checklist
    checklist = get_checklist(_base_module(module))
    done = total = pct = 0
    if checklist:
        states = get_checklist_states(db, module)
        custom = get_custom_tasks(db, module)
        hidden = get_hidden_tasks(db, module)
        done, total, pct = checklist_progress(checklist, states, custom, hidden)
    return jsonify({"ok": True, "done": done, "total": total, "percent": pct})


@app.route("/checklist/reset/<module>", methods=["POST"])
@login_required
def checklist_reset(module):
    db = get_db()
    db.execute("DELETE FROM checklist_state WHERE module=?", (module,))
    db.commit()
    log_action("task_reset", "task",
               "Reset all checklist progress for %s" % friendly_scope(db, module), target=module)
    flash("Checklist reset (custom tasks kept).", "info")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/checklist/task/add", methods=["POST"])
@login_required
def checklist_task_add():
    db = get_db()
    data = request.get_json(silent=True) or request.form
    module = (data.get("module") or "").strip()
    step_key = (data.get("step_key") or "").strip()
    text = (data.get("text") or "").strip()
    if not module or not step_key or not text:
        return jsonify({"ok": False, "error": "module, step_key and text are required"}), 400

    cur = db.execute(
        "INSERT INTO custom_tasks (module, step_key, text) VALUES (?,?,?)",
        (module, step_key, text),
    )
    db.commit()
    new_id = cur.lastrowid
    log_action("task_add", "task",
               "Added custom task '%s' in %s" % (text, friendly_scope(db, module)), target=module)

    from modules.checklists import get_checklist
    checklist = get_checklist(_base_module(module))
    done = total = pct = 0
    if checklist:
        states = get_checklist_states(db, module)
        custom = get_custom_tasks(db, module)
        hidden = get_hidden_tasks(db, module)
        done, total, pct = checklist_progress(checklist, states, custom, hidden)

    return jsonify({
        "ok": True,
        "task": {"id": new_id, "text": text, "key": "custom-%d" % new_id},
        "done": done, "total": total, "percent": pct,
    })


@app.route("/checklist/task/delete", methods=["POST"])
@login_required
def checklist_task_delete():
    db = get_db()
    data = request.get_json(silent=True) or request.form
    task_id = data.get("task_id")
    module = (data.get("module") or "").strip()
    if not task_id:
        return jsonify({"ok": False, "error": "task_id required"}), 400

    deleted_text = resolve_item_text(db, module, "custom-%s" % task_id)
    db.execute("DELETE FROM custom_tasks WHERE id=? AND module=?", (task_id, module))
    db.execute(
        "DELETE FROM checklist_state WHERE module=? AND item_key=?",
        (module, "custom-%s" % task_id),
    )
    db.commit()
    log_action("task_delete", "task",
               "Deleted custom task '%s' in %s" % (deleted_text, friendly_scope(db, module)),
               target=module)

    from modules.checklists import get_checklist
    checklist = get_checklist(_base_module(module))
    done = total = pct = 0
    if checklist:
        states = get_checklist_states(db, module)
        custom = get_custom_tasks(db, module)
        hidden = get_hidden_tasks(db, module)
        done, total, pct = checklist_progress(checklist, states, custom, hidden)

    return jsonify({"ok": True, "done": done, "total": total, "percent": pct})


@app.route("/checklist/task/hide", methods=["POST"])
@login_required
def checklist_task_hide():
    """Hide a default (predefined) item from this module/scope."""
    db = get_db()
    data = request.get_json(silent=True) or request.form
    module = (data.get("module") or "").strip()
    item_key = (data.get("item_key") or "").strip()
    if not module or not item_key:
        return jsonify({"ok": False, "error": "module and item_key required"}), 400

    item_text = resolve_item_text(db, module, item_key)
    db.execute(
        """INSERT INTO hidden_tasks (module, item_key) VALUES (?, ?)
           ON CONFLICT(module, item_key) DO NOTHING""",
        (module, item_key),
    )
    # Also clear any checkbox state for this hidden item
    db.execute(
        "DELETE FROM checklist_state WHERE module=? AND item_key=?",
        (module, item_key),
    )
    db.commit()
    log_action("task_hide", "task",
               "Removed default task '%s' from %s" % (item_text, friendly_scope(db, module)),
               target=module)

    from modules.checklists import get_checklist
    checklist = get_checklist(_base_module(module))
    done = total = pct = 0
    if checklist:
        states = get_checklist_states(db, module)
        custom = get_custom_tasks(db, module)
        hidden = get_hidden_tasks(db, module)
        done, total, pct = checklist_progress(checklist, states, custom, hidden)

    return jsonify({"ok": True, "done": done, "total": total, "percent": pct})


@app.route("/checklist/restore-defaults/<path:module>", methods=["POST"])
@login_required
def checklist_restore_defaults(module):
    """Restore all hidden default items for a module/scope."""
    db = get_db()
    db.execute("DELETE FROM hidden_tasks WHERE module=?", (module,))
    db.commit()
    log_action("task_restore_defaults", "task",
               "Restored hidden default tasks for %s" % friendly_scope(db, module), target=module)
    flash("Default tasks restored.", "info")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/api/bbbee/level", methods=["POST"])
@login_required
def api_bbbee_level():
    from modules.bbbee import calculate_level
    data = request.get_json()
    scorecard_type = data.get("scorecard_type", "Generic")
    total = float(data.get("total_score", 0))
    black_pct = float(data.get("black_owned_percent", 0))
    level, rating = calculate_level(scorecard_type, total, black_pct)
    return jsonify({"level": level, "rating": rating})


# ─── Audit log (admin only) ──────────────────────────────────────────────────

AUDIT_CATEGORIES = [
    ("auth", "Authentication"),
    ("project", "Projects"),
    ("task", "Tasks"),
    ("trash", "Trash bin"),
    ("user", "User management"),
    ("deadline", "Deadlines"),
]


@app.route("/audit", methods=["GET"])
@admin_required
def audit():
    db = get_db()

    # Filters
    f_user = request.args.get("user", "").strip()
    f_category = request.args.get("category", "").strip()
    f_search = request.args.get("q", "").strip()
    f_from = request.args.get("from", "").strip()
    f_to = request.args.get("to", "").strip()

    where, params = [], []
    if f_user:
        where.append("username = ?")
        params.append(f_user)
    if f_category:
        where.append("category = ?")
        params.append(f_category)
    if f_search:
        where.append("(detail LIKE ? OR action LIKE ? OR target LIKE ?)")
        like = "%" + f_search + "%"
        params += [like, like, like]
    if f_from:
        where.append("date(ts) >= date(?)")
        params.append(f_from)
    if f_to:
        where.append("date(ts) <= date(?)")
        params.append(f_to)

    sql = "SELECT * FROM audit_log"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC LIMIT 1000"
    entries = db.execute(sql, params).fetchall()

    # Distinct users for the filter dropdown
    users = [r["username"] for r in db.execute(
        "SELECT DISTINCT username FROM audit_log ORDER BY username"
    ).fetchall()]

    total_entries = db.execute("SELECT COUNT(*) AS c FROM audit_log").fetchone()["c"]

    return render_template(
        "audit.html",
        entries=entries,
        users=users,
        categories=AUDIT_CATEGORIES,
        total_entries=total_entries,
        shown=len(entries),
        f_user=f_user, f_category=f_category, f_search=f_search,
        f_from=f_from, f_to=f_to,
        any_filter=bool(where),
    )


# ─── User management (admin only) ────────────────────────────────────────────

@app.route("/users", methods=["GET"])
@admin_required
def users_admin():
    db = get_db()
    rows = db.execute("SELECT * FROM users ORDER BY role, username").fetchall()
    return render_template("users.html", users=rows)


@app.route("/users/add", methods=["POST"])
@admin_required
def users_add():
    db = get_db()
    username = (request.form.get("username") or "").strip()
    full_name = (request.form.get("full_name") or "").strip()
    password = request.form.get("password") or ""
    role = request.form.get("role") or "user"
    if role not in ("admin", "user"):
        role = "user"
    if not username or not password:
        flash("Username and password are required.", "danger")
        return redirect(url_for("users_admin"))
    existing = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if existing:
        flash("That username already exists.", "danger")
        return redirect(url_for("users_admin"))
    db.execute(
        "INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)",
        (username, generate_password_hash(password), role, full_name or username),
    )
    db.commit()
    log_action("user_create", "user",
               f"Created {role} account '{username}'", target=username)
    flash(f"User '{username}' created ({role}).", "success")
    return redirect(url_for("users_admin"))


@app.route("/users/<int:user_id>/reset", methods=["POST"])
@admin_required
def users_reset(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("users_admin"))
    new_password = request.form.get("password") or ""
    if not new_password:
        flash("Please provide a new password.", "danger")
        return redirect(url_for("users_admin"))
    db.execute("UPDATE users SET password_hash=? WHERE id=?",
               (generate_password_hash(new_password), user_id))
    db.commit()
    log_action("user_reset_password", "user",
               f"Reset password for '{user['username']}'", target=user["username"])
    flash(f"Password reset for '{user['username']}'.", "success")
    return redirect(url_for("users_admin"))


@app.route("/users/<int:user_id>/toggle-role", methods=["POST"])
@admin_required
def users_toggle_role(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("users_admin"))
    new_role = "user" if user["role"] == "admin" else "admin"
    # Guard: don't allow removing the last admin
    admins = db.execute("SELECT COUNT(*) AS c FROM users WHERE role='admin'").fetchone()["c"]
    if user["role"] == "admin" and admins <= 1:
        flash("Cannot demote the last administrator.", "danger")
        return redirect(url_for("users_admin"))
    db.execute("UPDATE users SET role=? WHERE id=?", (new_role, user_id))
    db.commit()
    log_action("user_role_change", "user",
               f"Changed role of '{user['username']}' to {new_role}", target=user["username"])
    flash(f"'{user['username']}' is now {new_role}.", "success")
    return redirect(url_for("users_admin"))


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def users_delete(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("users_admin"))
    if user["username"] == session.get("username"):
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("users_admin"))
    if user["role"] == "admin":
        admins = db.execute("SELECT COUNT(*) AS c FROM users WHERE role='admin'").fetchone()["c"]
        if admins <= 1:
            flash("Cannot delete the last administrator.", "danger")
            return redirect(url_for("users_admin"))
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit()
    log_action("user_delete", "user", f"Deleted account '{user['username']}'", target=user["username"])
    flash(f"User '{user['username']}' deleted.", "warning")
    return redirect(url_for("users_admin"))


# ── Template context processor ───────────────────────────────

@app.context_processor
def inject_now():
    return {
        "now": datetime.now,
        "current_user": session.get("username", ""),
        "current_full_name": session.get("full_name", session.get("username", "")),
        "current_role": session.get("role", ""),
        "is_admin": session.get("role") == "admin",
        "session_remember": session.get("remember", False),
        "inactivity_timeout": INACTIVITY_TIMEOUT,
    }


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
