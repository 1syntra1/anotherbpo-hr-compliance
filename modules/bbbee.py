"""BBBEE scorecard calculation and report generation."""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    RL_AVAILABLE = True
except ImportError:
    RL_AVAILABLE = False

NAVY = "1A2744"
LIGHT_BLUE = "D9E8F5"
HEADER_BLUE = "2D9CDB"
WHITE = "FFFFFF"
GREY = "F2F2F2"

# BBBEE Level thresholds (Generic 109-point scorecard)
GENERIC_THRESHOLDS = [
    (100, 1), (95, 2), (90, 3), (80, 4), (75, 5), (65, 6), (55, 7), (0, 8)
]

# QSE thresholds (100-point)
QSE_THRESHOLDS = [
    (100, 1), (90, 2), (80, 3), (70, 4), (60, 5), (50, 6), (40, 7), (0, 8)
]

LEVEL_COLOURS = {
    1: "#006400",  # Dark green
    2: "#228B22",
    3: "#32CD32",
    4: "#9ACD32",
    5: "#FFD700",
    6: "#FFA500",
    7: "#FF6347",
    8: "#FF0000",  # Red
}

LEVEL_RECOGNITION = {
    1: "135%", 2: "125%", 3: "110%", 4: "100%",
    5: "80%", 6: "60%", 7: "50%", 8: "10%",
    "Non-Compliant": "0%",
}


def calculate_level(scorecard_type: str, total_score: float, black_owned_percent: float = 0) -> tuple:
    """Return (level, recognition_percentage) for a given scorecard and score."""
    stype = scorecard_type.upper()

    # EME: automatic level
    if stype == "EME":
        if black_owned_percent >= 51:
            level = 1
        else:
            level = 4
        return level, LEVEL_RECOGNITION.get(level, "100%")

    thresholds = GENERIC_THRESHOLDS if stype == "GENERIC" else QSE_THRESHOLDS
    for threshold, lvl in thresholds:
        if total_score >= threshold:
            return lvl, LEVEL_RECOGNITION.get(lvl, "0%")
    return 8, LEVEL_RECOGNITION[8]


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def generate_scorecard_excel(scorecard_data: dict) -> bytes:
    """Generate BBBEE scorecard Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BBBEE Scorecard"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    border = _thin_border()

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "BROAD-BASED BLACK ECONOMIC EMPOWERMENT (BBBEE) SCORECARD"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = _hfill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Company info
    info = [
        ("Company:", scorecard_data.get("company_name", "")),
        ("Assessment Year:", scorecard_data.get("assessment_year", str(datetime.now().year))),
        ("Scorecard Type:", scorecard_data.get("scorecard_type", "Generic")),
        ("BBBEE Level:", scorecard_data.get("bbbee_level", 8)),
        ("Generated:", datetime.now().strftime("%d %B %Y")),
    ]
    for i, (label, value) in enumerate(info, 2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"A{i}"].fill = _hfill(LIGHT_BLUE)
        ws[f"A{i}"].border = border
        ws[f"B{i}"].border = border

    # Scorecard table
    stype = scorecard_data.get("scorecard_type", "Generic").upper()
    header_row = 8
    headers = ["Element", "Max Points", "Score Achieved", "% Achievement", "Weighting"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = _hfill(HEADER_BLUE)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 28

    if stype == "GENERIC":
        elements = [
            ("Ownership", 25, scorecard_data.get("ownership_score", 0)),
            ("Management Control", 19, scorecard_data.get("management_score", 0)),
            ("Skills Development", 20, scorecard_data.get("skills_score", 0)),
            ("Enterprise & Supplier Development", 40, scorecard_data.get("esd_score", 0)),
            ("Socio-Economic Development", 5, scorecard_data.get("sed_score", 0)),
        ]
        max_total = 109
    elif stype == "QSE":
        elements = [
            ("Ownership", 25, scorecard_data.get("ownership_score", 0)),
            ("Management Control", 25, scorecard_data.get("management_score", 0)),
            ("Skills Development", 25, scorecard_data.get("skills_score", 0)),
            ("Enterprise & Supplier Development", 25, scorecard_data.get("esd_score", 0)),
        ]
        max_total = 100
    else:  # EME
        elements = [
            ("Automatic BBBEE Level (EME)", 100, 100),
        ]
        max_total = 100

    for i, (element, max_pts, score) in enumerate(elements):
        row = header_row + 1 + i
        fill = _hfill(GREY) if i % 2 == 0 else PatternFill()
        pct = (score / max_pts * 100) if max_pts else 0
        weight = (max_pts / max_total * 100) if max_total else 0
        vals = [element, max_pts, score, f"{pct:.1f}%", f"{weight:.1f}%"]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.fill = fill
            if col >= 2:
                c.alignment = Alignment(horizontal="center")

    # Total row
    total_row = header_row + 1 + len(elements)
    total_score = scorecard_data.get("total_score", 0)
    ws.cell(row=total_row, column=1, value="TOTAL SCORE").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = _hfill(NAVY)
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=2, value=max_total).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=2).fill = _hfill(NAVY)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=2).border = border
    ws.cell(row=total_row, column=3, value=total_score).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=3).fill = _hfill(NAVY)
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=3).border = border

    level = scorecard_data.get("bbbee_level", 8)
    recognition = LEVEL_RECOGNITION.get(level, "0%")

    level_row = total_row + 2
    ws.merge_cells(f"A{level_row}:B{level_row}")
    ws.cell(row=level_row, column=1, value=f"BBBEE LEVEL: {level}").font = Font(bold=True, size=14, color=WHITE)
    ws.cell(row=level_row, column=1).fill = _hfill(NAVY)
    ws.cell(row=level_row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(f"C{level_row}:E{level_row}")
    ws.cell(row=level_row, column=3, value=f"Procurement Recognition: {recognition}").font = Font(bold=True)
    ws.cell(row=level_row, column=3).fill = _hfill(LIGHT_BLUE)
    ws.cell(row=level_row, column=3).alignment = Alignment(horizontal="center")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_scorecard_pdf(scorecard_data: dict) -> bytes:
    """Generate BBBEE Certificate Summary PDF."""
    if not RL_AVAILABLE:
        raise RuntimeError("reportlab not installed.")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                             leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    navy_rl = colors.HexColor("#1A2744")
    blue_rl = colors.HexColor("#2D9CDB")
    light_blue_rl = colors.HexColor("#D9E8F5")

    # Title
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        textColor=colors.white, backColor=navy_rl,
        fontSize=16, spaceAfter=0, alignment=TA_CENTER,
        borderPad=10,
    )
    story.append(Paragraph("BBBEE COMPLIANCE CERTIFICATE SUMMARY", title_style))
    story.append(Spacer(1, 0.5*cm))

    # Sub-heading
    story.append(Paragraph(
        "Broad-Based Black Economic Empowerment — Scorecard Summary",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=11, alignment=TA_CENTER,
                       textColor=navy_rl)
    ))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=blue_rl))
    story.append(Spacer(1, 0.3*cm))

    # Company info table
    level = scorecard_data.get("bbbee_level", 8)
    recognition = LEVEL_RECOGNITION.get(level, "0%")
    stype = scorecard_data.get("scorecard_type", "Generic")

    info_data = [
        ["Company Name:", scorecard_data.get("company_name", "N/A")],
        ["Assessment Year:", scorecard_data.get("assessment_year", str(datetime.now().year))],
        ["Scorecard Type:", stype],
        ["Total Score:", f"{scorecard_data.get('total_score', 0):.2f}"],
        ["BBBEE Level:", f"Level {level}"],
        ["Procurement Recognition:", recognition],
        ["Report Date:", datetime.now().strftime("%d %B %Y")],
    ]
    info_table = Table(info_data, colWidths=[6*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), light_blue_rl),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (1, 0), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    # Scorecard breakdown
    story.append(Paragraph("Scorecard Breakdown", ParagraphStyle(
        "h2", parent=styles["Heading2"], textColor=navy_rl
    )))

    if stype.upper() == "GENERIC":
        elements = [
            ["Element", "Max Points", "Score"],
            ["Ownership", "25", f"{scorecard_data.get('ownership_score', 0):.2f}"],
            ["Management Control", "19", f"{scorecard_data.get('management_score', 0):.2f}"],
            ["Skills Development", "20", f"{scorecard_data.get('skills_score', 0):.2f}"],
            ["Enterprise & Supplier Development", "40", f"{scorecard_data.get('esd_score', 0):.2f}"],
            ["Socio-Economic Development", "5", f"{scorecard_data.get('sed_score', 0):.2f}"],
            ["TOTAL", "109", f"{scorecard_data.get('total_score', 0):.2f}"],
        ]
    elif stype.upper() == "QSE":
        elements = [
            ["Element", "Max Points", "Score"],
            ["Ownership", "25", f"{scorecard_data.get('ownership_score', 0):.2f}"],
            ["Management Control", "25", f"{scorecard_data.get('management_score', 0):.2f}"],
            ["Skills Development", "25", f"{scorecard_data.get('skills_score', 0):.2f}"],
            ["Enterprise & Supplier Development", "25", f"{scorecard_data.get('esd_score', 0):.2f}"],
            ["TOTAL", "100", f"{scorecard_data.get('total_score', 0):.2f}"],
        ]
    else:
        elements = [
            ["Element", "Details"],
            ["Scorecard Type", "Exempted Micro Enterprise (EME)"],
            ["BBBEE Level", f"Level {level} (Automatic)"],
            ["Recognition", recognition],
        ]

    score_table = Table(elements, colWidths=[9*cm, 3.5*cm, 3.5*cm] if len(elements[0]) == 3 else [9*cm, 7*cm])
    score_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy_rl),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), light_blue_rl),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(score_table)
    story.append(Spacer(1, 0.5*cm))

    # Level badge
    level_color = colors.HexColor(LEVEL_COLOURS.get(level, "#FF0000"))
    level_data = [[f"BBBEE LEVEL {level}", f"Procurement Recognition: {recognition}"]]
    level_table = Table(level_data, colWidths=[8*cm, 8*cm])
    level_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), level_color),
        ("BACKGROUND", (1, 0), (1, 0), light_blue_rl),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("GRID", (0, 0), (-1, -1), 1, colors.white),
    ]))
    story.append(level_table)
    story.append(Spacer(1, 0.5*cm))

    # Disclaimer
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "This certificate summary is generated by the HR Audit System for internal use. "
        "Official BBBEE verification must be conducted by an accredited verification agency.",
        ParagraphStyle("disc", parent=styles["Normal"], fontSize=8, textColor=colors.grey,
                       alignment=TA_CENTER)
    ))

    doc.build(story)
    return buf.getvalue()
