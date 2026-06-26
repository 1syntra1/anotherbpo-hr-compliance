"""COIDA W.As.8 Return of Earnings report generation."""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1A2744"
LIGHT_BLUE = "D9E8F5"
HEADER_BLUE = "2D9CDB"
WHITE = "FFFFFF"
GREY = "F2F2F2"

# COIDA assessment rates per class (approximate 2024 rates per R100 earnings)
# In practice these vary by class. These are indicative rates.
ASSESSMENT_RATES = {
    "A": 0.65,   # Administration
    "B": 0.71,
    "C": 0.80,
    "D": 1.20,
    "E": 1.45,
    "F": 1.90,
    "G": 2.30,
    "H": 3.10,
    "I": 4.50,
    "J": 6.00,
    "K": 8.50,
    "L": 12.00,
    "M": 18.00,
    "N": 25.00,
    "O": 35.00,
    "DEFAULT": 1.00,
}

# Maximum earnings per employee (2024/2025 ceiling)
MAX_EARNINGS_PER_EMPLOYEE = 563520.00


def calculate_assessment(earnings_records) -> dict:
    """
    Calculate COIDA levy per class and return summary dict.
    earnings_records: list of sqlite3.Row or dict with keys:
        class_code, subclass, annual_earnings, num_employees
    """
    total_earnings = 0.0
    total_levy = 0.0
    details = []

    for rec in earnings_records:
        if isinstance(rec, dict):
            class_code = rec.get("class_code", "DEFAULT") or "DEFAULT"
            annual_earnings = float(rec.get("annual_earnings") or 0)
            num_employees = int(rec.get("num_employees") or 0)
            subclass = rec.get("subclass", "")
            description = rec.get("description", "")
        else:
            class_code = rec["class_code"] or "DEFAULT"
            annual_earnings = float(rec["annual_earnings"] or 0)
            num_employees = int(rec["num_employees"] or 0)
            subclass = rec["subclass"] or ""
            description = rec["description"] or ""

        # Cap earnings per employee
        capped_per_emp = min(annual_earnings / num_employees, MAX_EARNINGS_PER_EMPLOYEE) if num_employees else 0
        capped_total = capped_per_emp * num_employees

        rate = ASSESSMENT_RATES.get(class_code.upper(), ASSESSMENT_RATES["DEFAULT"])
        levy = (capped_total / 100) * rate

        total_earnings += capped_total
        total_levy += levy
        details.append({
            "class_code": class_code,
            "subclass": subclass,
            "description": description,
            "num_employees": num_employees,
            "annual_earnings": annual_earnings,
            "capped_earnings": round(capped_total, 2),
            "rate": rate,
            "levy": round(levy, 2),
        })

    return {
        "details": details,
        "total_earnings": round(total_earnings, 2),
        "total_levy": round(total_levy, 2),
        "max_earnings_per_employee": MAX_EARNINGS_PER_EMPLOYEE,
    }


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def generate_was8(db, company_info: dict) -> bytes:
    """Generate W.As.8 Return of Earnings Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "W.As.8"

    col_widths = [6, 12, 22, 20, 16, 22, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    border = _thin_border()

    # Main title block
    ws.merge_cells("A1:H1")
    ws["A1"] = "COMPENSATION FUND — W.As.8 RETURN OF EARNINGS"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = _hfill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = "Compensation for Occupational Injuries and Diseases Act, 130 of 1993"
    ws["A2"].font = Font(italic=True, size=10, color=WHITE)
    ws["A2"].fill = _hfill(HEADER_BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Employer details
    info_pairs = [
        ("Employer Reg. No.:", company_info.get("employer_reg_number", "")),
        ("Company Name:", company_info.get("company_name", "")),
        ("Assessment Year:", company_info.get("assessment_year", str(datetime.now().year))),
        ("Report Generated:", datetime.now().strftime("%d %B %Y")),
    ]
    for i, (label, value) in enumerate(info_pairs, 4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws.merge_cells(f"B{i}:D{i}")
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"A{i}"].fill = _hfill(LIGHT_BLUE)
        ws[f"A{i}"].border = border
        ws[f"B{i}"].border = border

    # Table headers
    header_row = 9
    headers = [
        "#", "Class Code", "Subclass", "Description",
        "No. Employees", "Annual Earnings (R)", "Capped Earnings (R)",
        "Estimated Levy (R)"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = _hfill(NAVY)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 36

    # Note about cap
    ws.cell(row=8, column=1,
            value=f"Note: Annual earnings capped at R{MAX_EARNINGS_PER_EMPLOYEE:,.2f} per employee as per COIDA regulations.")
    ws.cell(row=8, column=1).font = Font(italic=True, size=9, color="666666")
    ws.merge_cells("A8:H8")

    # Data rows
    records = db.execute("SELECT * FROM coida_classes ORDER BY class_code").fetchall()
    assessment = calculate_assessment(list(records))

    for i, detail in enumerate(assessment["details"]):
        row = header_row + 1 + i
        fill = _hfill(GREY) if i % 2 == 0 else PatternFill()
        values = [
            i + 1,
            detail["class_code"],
            detail["subclass"],
            detail["description"],
            detail["num_employees"],
            detail["annual_earnings"],
            detail["capped_earnings"],
            detail["levy"],
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.fill = fill
            if col in [5, 6, 7, 8]:
                c.alignment = Alignment(horizontal="right")
            if col in [6, 7, 8]:
                c.number_format = "#,##0.00"

    # Totals row
    total_row = header_row + 1 + len(assessment["details"])
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = _hfill(NAVY)
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=total_row, column=6, value=assessment["total_earnings"]).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=6).fill = _hfill(NAVY)
    ws.cell(row=total_row, column=6).number_format = "#,##0.00"
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=6).border = border

    ws.cell(row=total_row, column=7, value="").fill = _hfill(NAVY)
    ws.cell(row=total_row, column=7).border = border

    ws.cell(row=total_row, column=8, value=assessment["total_levy"]).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=8).fill = _hfill(NAVY)
    ws.cell(row=total_row, column=8).number_format = "#,##0.00"
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=8).border = border

    # Summary box
    summary_row = total_row + 2
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    ws.cell(row=summary_row, column=1, value="ASSESSMENT SUMMARY").font = Font(bold=True, size=12, color=WHITE)
    ws.cell(row=summary_row, column=1).fill = _hfill(NAVY)
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center")

    summary_pairs = [
        ("Total Gross Earnings:", assessment["total_earnings"]),
        ("Total Estimated Assessment (Levy):", assessment["total_levy"]),
        ("Max Earnings Per Employee:", MAX_EARNINGS_PER_EMPLOYEE),
    ]
    for i, (label, value) in enumerate(summary_pairs, summary_row + 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = _hfill(LIGHT_BLUE)
        ws.merge_cells(f"A{i}:C{i}")
        ws.cell(row=i, column=4, value=value).number_format = "#,##0.00"
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=4).font = Font(bold=True)
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = border

    # Disclaimer
    disclaimer_row = summary_row + len(summary_pairs) + 2
    ws.merge_cells(f"A{disclaimer_row}:H{disclaimer_row}")
    ws.cell(row=disclaimer_row, column=1,
            value="DISCLAIMER: This is a preliminary estimate. Final assessment by the Compensation Fund may differ. "
                  "Submit official W.As.8 form at www.cf.gov.za")
    ws.cell(row=disclaimer_row, column=1).font = Font(italic=True, size=9, color="666666")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
