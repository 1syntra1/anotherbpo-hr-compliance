"""GBS / Employment Equity report generation (EEA2 and EEA4)."""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

NAVY = "1A2744"
LIGHT_BLUE = "D9E8F5"
HEADER_BLUE = "2D9CDB"
WHITE = "FFFFFF"
GREY = "F2F2F2"

OCC_LEVELS = [
    "Top Management", "Senior Management", "Professionally Qualified",
    "Skilled Technical", "Semi-Skilled", "Unskilled", "Non-Permanent"
]
RACES = ["African", "Coloured", "Indian", "White", "Foreign National"]


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _build_matrix(db):
    employees = db.execute("SELECT * FROM employees").fetchall()
    matrix = {lvl: {r: {"M": 0, "F": 0} for r in RACES} for lvl in OCC_LEVELS}
    for emp in employees:
        lvl = emp["occupational_level"]
        race = emp["race"]
        gender = emp["gender"]
        if lvl in matrix and race in RACES:
            key = "M" if gender == "Male" else "F"
            matrix[lvl][race][key] += 1
    return matrix


def generate_eea2(db, company_info: dict) -> bytes:
    """Generate EEA2 Employment Equity Report (xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EEA2"

    # Set column widths
    ws.column_dimensions["A"].width = 28
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Title block
    ws.merge_cells("A1:M1")
    ws["A1"] = "EMPLOYMENT EQUITY ACT, 55 OF 1998"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = _header_fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = "EEA2 – EMPLOYMENT EQUITY REPORT"
    ws["A2"].font = Font(bold=True, size=12, color=WHITE)
    ws["A2"].fill = _header_fill(NAVY)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Company info block
    ws["A3"] = "Company Name:"
    ws["B3"] = company_info.get("company_name", "")
    ws["A4"] = "Registration Number:"
    ws["B4"] = company_info.get("reg_number", "")
    ws["A5"] = "Reporting Year:"
    ws["B5"] = company_info.get("reporting_year", str(datetime.now().year))
    ws["A6"] = "Designated Employer:"
    ws["B6"] = company_info.get("designated_employer", "Yes")
    ws["A7"] = "Report Generated:"
    ws["B7"] = datetime.now().strftime("%d %B %Y")
    for row in range(3, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = _header_fill(LIGHT_BLUE)

    # Matrix headers
    start_row = 9
    ws.merge_cells(f"A{start_row}:A{start_row+1}")
    ws[f"A{start_row}"] = "Occupational Level"
    ws[f"A{start_row}"].font = Font(bold=True, color=WHITE)
    ws[f"A{start_row}"].fill = _header_fill(NAVY)
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    race_cols = [
        ("African", 2), ("Coloured", 4), ("Indian", 6),
        ("White", 8), ("Foreign National", 10)
    ]
    col_offset = 2
    for race, col in race_cols:
        end_col = col + 1
        ws.merge_cells(
            start_row=start_row, start_column=col,
            end_row=start_row, end_column=end_col
        )
        cell = ws.cell(row=start_row, column=col, value=race)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _header_fill(HEADER_BLUE)
        cell.alignment = Alignment(horizontal="center")

        ws.cell(row=start_row + 1, column=col, value="Male").font = Font(bold=True)
        ws.cell(row=start_row + 1, column=col + 1, value="Female").font = Font(bold=True)
        for c in [col, col + 1]:
            ws.cell(row=start_row + 1, column=c).fill = _header_fill(LIGHT_BLUE)
            ws.cell(row=start_row + 1, column=c).alignment = Alignment(horizontal="center")

    # Total column
    ws.merge_cells(
        start_row=start_row, start_column=12,
        end_row=start_row, end_column=13
    )
    ws.cell(row=start_row, column=12, value="Total").font = Font(bold=True, color=WHITE)
    ws.cell(row=start_row, column=12).fill = _header_fill(NAVY)
    ws.cell(row=start_row, column=12).alignment = Alignment(horizontal="center")
    ws.cell(row=start_row + 1, column=12, value="Male").font = Font(bold=True)
    ws.cell(row=start_row + 1, column=13, value="Female").font = Font(bold=True)
    for c in [12, 13]:
        ws.cell(row=start_row + 1, column=c).fill = _header_fill(LIGHT_BLUE)
        ws.cell(row=start_row + 1, column=c).alignment = Alignment(horizontal="center")

    matrix = _build_matrix(db)
    data_start = start_row + 2
    border = _thin_border()

    for i, lvl in enumerate(OCC_LEVELS):
        row = data_start + i
        fill = _header_fill(GREY) if i % 2 == 0 else PatternFill()
        ws.cell(row=row, column=1, value=lvl).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = border

        total_m = total_f = 0
        for j, race in enumerate(RACES):
            col_m = 2 + j * 2
            col_f = col_m + 1
            m_val = matrix[lvl][race]["M"]
            f_val = matrix[lvl][race]["F"]
            ws.cell(row=row, column=col_m, value=m_val).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=col_f, value=f_val).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=col_m).border = border
            ws.cell(row=row, column=col_f).border = border
            ws.cell(row=row, column=col_m).fill = fill
            ws.cell(row=row, column=col_f).fill = fill
            total_m += m_val
            total_f += f_val

        ws.cell(row=row, column=12, value=total_m).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=13, value=total_f).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=12).border = border
        ws.cell(row=row, column=13).border = border
        ws.cell(row=row, column=12).font = Font(bold=True)
        ws.cell(row=row, column=13).font = Font(bold=True)

    # Grand total row
    grand_row = data_start + len(OCC_LEVELS)
    ws.cell(row=grand_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=grand_row, column=1).fill = _header_fill(NAVY)
    for col in range(2, 14):
        col_letter = get_column_letter(col)
        formula_start = get_column_letter(col) + str(data_start)
        formula_end = get_column_letter(col) + str(grand_row - 1)
        ws.cell(row=grand_row, column=col).value = f"=SUM({formula_start}:{formula_end})"
        ws.cell(row=grand_row, column=col).font = Font(bold=True, color=WHITE)
        ws.cell(row=grand_row, column=col).fill = _header_fill(NAVY)
        ws.cell(row=grand_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=grand_row, column=col).border = border

    # Disability note
    disability_row = grand_row + 2
    emp_count = db.execute("SELECT COUNT(*) as c FROM employees WHERE disability='Yes'").fetchone()
    ws.cell(row=disability_row, column=1, value="People with Disabilities:").font = Font(bold=True)
    ws.cell(row=disability_row, column=2, value=emp_count["c"] if emp_count else 0)
    ws.cell(row=disability_row, column=1).fill = _header_fill(LIGHT_BLUE)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_eea4(db, company_info: dict) -> bytes:
    """Generate EEA4 Income Differential Report (xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EEA4"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "EEA4 – INCOME DIFFERENTIAL STATEMENT"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = _header_fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Company:"
    ws["B2"] = company_info.get("company_name", "")
    ws["A3"] = "Reporting Year:"
    ws["B3"] = company_info.get("reporting_year", str(datetime.now().year))
    ws["A4"] = "Generated:"
    ws["B4"] = datetime.now().strftime("%d %B %Y")
    for r in range(2, 5):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].fill = _header_fill(LIGHT_BLUE)

    headers = ["Occupational Level", "Race", "Gender", "Headcount", "Total Annual Earnings (R)", "Average Salary (R)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = _header_fill(HEADER_BLUE)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[6].height = 30

    employees = db.execute("SELECT * FROM employees").fetchall()
    row_num = 7
    border = _thin_border()
    for lvl in OCC_LEVELS:
        for race in RACES:
            for gender in ["Male", "Female"]:
                subset = [
                    e for e in employees
                    if e["occupational_level"] == lvl
                    and e["race"] == race
                    and e["gender"] == gender
                ]
                if not subset:
                    continue
                headcount = len(subset)
                total_earn = sum((e["salary"] or 0) * 12 for e in subset)
                avg = total_earn / headcount if headcount else 0
                fill = _header_fill(GREY) if row_num % 2 == 0 else PatternFill()
                values = [lvl, race, gender, headcount, round(total_earn, 2), round(avg, 2)]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.border = border
                    c.fill = fill
                    if col >= 4:
                        c.alignment = Alignment(horizontal="right")
                        if col >= 5:
                            c.number_format = "#,##0.00"
                row_num += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
